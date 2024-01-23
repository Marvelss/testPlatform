"""
@Author  : SakuraFox
@Time    : 2024-01-23 10:35
@File    : demo39.py
@Description : 提取图片(成功)
"""
import math

from openpyxl import load_workbook
import os
from PIL import Image

# 命名规则
good_code = ""
# 命名字典
name_dict = {}
# 图片数量
img_num = 0
# 记录哪些文件夹已经被提取过了
folder_name_dict = {}
# 是否输出提取文本
is_text = True


# 创建文件夹
def create_folder():
    if not os.path.exists('images'):
        os.makedirs('images')
    print("成功创建/更新images文件夹!")


# 提取图片
def extract_images(stop):
    if stop:
        return
    global img_num
    global good_code
    folder = input("请输入需要提取的文件夹名称（不输入则遍历当前目录下未提取过的所有文件夹）：")
    # good_code = input("请输入命名规则对应表格中的名字（不输入则默认为货品编码）：")
    # if good_code == "":
    #     good_code = "货品编码"
    if folder != '':
        # 查找指定文件夹
        extract_images_from_excel(folder)
    else:
        folders = []
        for entry in os.scandir('..'):
            if entry.is_dir():
                folders.append(entry.name)
        i = 1
        # 记录可提取的文件夹的数量
        number = 0
        # 遍历当前目录下的文件夹
        for folder in folders:
            if folder in folder_name_dict:
                continue
            print(f"正在遍历第{i}个文件夹{folder}......")
            # 进行提取图片
            extract_images_from_excel(folder)
            number += 1
            i += 1
        if number == 0:
            print("没有可供提取的文件夹了！")
            return
    is_success()
    img_num = 0
    status = input("\n是否继续提取（输入Y表示是，输入其他则退出）：")
    if status == "Y" or status == "y":
        extract_images(False)
    else:
        extract_images(True)


def is_success():
    if img_num == 0:
        if is_text:
            print(f'没有提取到图片！')
    else:
        print(f'成功提取{img_num}张图片！')
        print("图片提取完成，请到images文件夹中查看！")


# 进行提取图片
def extract_images_from_excel(folder):
    global img_num
    global is_text
    is_have_excel = False
    path = os.path.join('..', folder)
    if not os.path.exists(path):
        print(f'{folder}文件夹未找到！')
        return
    # 判断文件夹是否已经被提取过了
    if folder not in folder_name_dict:
        is_text = True
    else:
        print(f'{folder}文件夹已经被提取过了！')
        is_text = False
        return
    try:
        # 遍历当前文件夹内的所有文件
        for entry in os.scandir(path):
            # 如果当前对象是文件且后缀是xlsx或者xls
            if entry.is_file() and entry.name.endswith('.xlsx'):
                is_have_excel = True
                print(f'{folder}下的Excel文件路径为：{entry.path}')
                # 打开当前文件
                wb = load_workbook(entry.path)
                # 遍历每一个Sheet
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # 获取当前列名为货品编码的列序号
                    code_index = ""
                    for column in sheet.iter_cols():
                        if column[0].value == good_code:
                            code_index = column[0].column
                            break
                    if code_index == "":
                        print(f'列名{good_code}在{entry.path}的文件中不存在！')
                        break
                    else:
                        folder_name_dict[folder] = True
                    # 遍历Sheet中的所有图片
                    for image in sheet._images:
                        # 获取图片中心行数，判断货品编码是哪一个
                        row_index = (int(((image.anchor._from.row + 1) + (image.anchor.to.row + 1)) / 2))
                        # 获取当前行的货品编码列的值(取中间值)
                        code = ""
                        if code_index != "":
                            code = str(sheet.cell(row=row_index, column=code_index).value)
                        # 获取图片格式
                        img_format = image.format
                        # 这个if else只是命名规则，不重要
                        if code not in name_dict:
                            name_dict[code] = 1
                        else:
                            name_dict[code] = name_dict[code] + 1
                        save_path = f"./images/{code}-{name_dict[code]}.{img_format}"
                        # 保存
                        file = open(save_path, "wb")
                        file.write(image.ref.getvalue())
                        file.close()
                        # 压缩图片
                        compress_and_save_image(save_path)
                        img_num += 1
                    break
                # 关闭文件对象
                wb.close()
    except FileNotFoundError:
        # 处理文件未找到异常
        print(f'{folder}文件夹未找到！')
        extract_images(good_code)
    except Exception as e:
        # 处理其他异常
        print("提取图片异常:", e)
    if not is_have_excel:
        print(f'{folder}文件夹内未找到Excel文件！')
        folder_name_dict[folder] = True


# 压缩图片
def compress_and_save_image(image_path):
    # 打开原始图片
    original_image = Image.open(image_path)
    # 检查文件大小，并根据需要进行进一步压缩，压缩到1M
    if os.path.getsize(image_path) > 1024 * 1024:
        size = os.path.getsize(image_path)
        # 压缩到1mb需要压缩的比例(百分比)
        quality = math.floor(((1024 * 1024) / size) * 100)
        original_image.save(image_path, optimize=True, quality=quality)
    original_image.close()


# v1.0：此版本是针对于图片分布不规则的情况，提取图片速度尚且较慢
# v1.1：此版本是针对于图片集中分布在一列的情况，能更快提取图片出来。
# v1.2：此版本解决图片位于Excel边界时存在的问题，只要图片中心行在这一行，就可以匹配相应的国家编码，同时不用去遍历，直接获取图片。
# v1.3：此版本是让用户自己输入指定的文件夹，增加异常交互。
# v1.4：此版本增加了对1MB以上图片的压缩，解决了多图片在同一单元格的问题。
if __name__ == '__main__':
    print("开始提取......")
    # 创建存放文件夹
    create_folder()
    # 提取图片
    extract_images(False)
    # 最后加入输入语句，以阻塞程序的执行
    input("按下任意键以关闭程序")
    # import office

    e_path = r'E:\a_python\program\testPlatform\test\已打商标数据14002-14501.xlsx'
    pdf_path = r'/test/已打商标数据14002-14501.pdf'
    image_path = r'E:\a_python\program\testPlatform\images'
    # office.excel.excel2pdf(excel_path=e_path,
    #                        pdf_path=pdf_path)
    # office.pdf.pdf2imgs(
    #     pdf_path=pdf_path,
    #     out_dir=image_path
    # )
    # import popdf

    # import excel2pdf

    # popdf.pdf2imgs(pdf_path, image_path, merge=True)
    # # poexcel.excel2pdf()
    # import pdf2docx

    # 将PDF文件转换成Word文档
    # pdf2docx.parse_images('test.pdf')